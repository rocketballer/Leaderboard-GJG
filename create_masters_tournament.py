import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Define file paths
source_file = "slates-green-jacket-picks-2025-04-10 (Leaderboard).xlsx"
output_file = "GJG_Masters_Tournament.xlsx"

# Try to read the source file to extract participant data
try:
    source_df = pd.read_excel(source_file)
    has_source = True
    print(f"Successfully read source file: {source_file}")
except Exception as e:
    has_source = False
    print(f"Could not read source file: {e}")
    print("Creating example workbook with sample data.")

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Rename the default worksheet to "CBS Leaderboard"
ws1 = wb.active
ws1.title = "CBS Leaderboard"

# Create the second tab "Participants Leaderboard"
ws2 = wb.create_sheet(title="Participants Leaderboard")

# Format Tab 1: CBS Sports-Style Leaderboard
# Define columns for Tab 1
columns_tab1 = ["POS", "PLAYER", "TO PAR", "THRU", "TODAY", "R1", "R2", "R3", "R4", "TOTAL", "CUT"]

# Set column headers
for col_idx, col_name in enumerate(columns_tab1, 1):
    ws1.cell(row=1, column=col_idx).value = col_name
    ws1.cell(row=1, column=col_idx).font = Font(bold=True)
    ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

# Adjust column widths for better visibility
ws1.column_dimensions['A'].width = 8  # POS
ws1.column_dimensions['B'].width = 25  # PLAYER
ws1.column_dimensions['C'].width = 10  # TO PAR
ws1.column_dimensions['D'].width = 8   # THRU
ws1.column_dimensions['E'].width = 10  # TODAY
ws1.column_dimensions['F'].width = 8   # R1
ws1.column_dimensions['G'].width = 8   # R2
ws1.column_dimensions['H'].width = 8   # R3
ws1.column_dimensions['I'].width = 8   # R4
ws1.column_dimensions['J'].width = 10  # TOTAL
ws1.column_dimensions['K'].width = 8   # CUT

# Add sample data (placeholder) for Tab 1
sample_data = [
    ["1", "Scottie Scheffler", "-10", "F", "-4", "69", "67", "69", "69", "274", ""],
    ["2", "Collin Morikawa", "-8", "F", "-2", "71", "67", "70", "68", "276", ""],
    ["T3", "Tommy Fleetwood", "-7", "F", "-3", "72", "68", "70", "67", "277", ""],
    ["T3", "Max Homa", "-7", "F", "-1", "70", "69", "69", "69", "277", ""],
    ["T5", "Ludvig Åberg", "-6", "F", "E", "67", "70", "70", "71", "278", ""],
    ["MC", "Jordan Spieth", "+4", "F", "+2", "74", "74", "", "", "148", "CUT"],
    ["MC", "Rory McIlroy", "+3", "F", "+1", "73", "74", "", "", "147", "CUT"],
    ["T7", "Xander Schauffele", "-5", "F", "-2", "70", "71", "69", "69", "279", ""],
    ["T7", "Patrick Cantlay", "-5", "F", "-3", "72", "70", "70", "67", "279", ""],
    ["T9", "Hideki Matsuyama", "-4", "F", "+1", "68", "70", "69", "73", "280", ""],
    ["T9", "Justin Thomas", "-4", "F", "-2", "73", "69", "70", "68", "280", ""],
    ["T11", "Jon Rahm", "-3", "F", "E", "72", "70", "68", "71", "281", ""]
]

# Add the sample data to Tab 1
for row_idx, row_data in enumerate(sample_data, 2):  # Start from row 2
    for col_idx, cell_value in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx).value = cell_value
        ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

# Apply conditional formatting for under/over par
red_fill = PatternFill(start_color="FFE6B8B7", end_color="FFE6B8B7", fill_type="solid")  # Over par (red)
green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")  # Under par (green)

# Create condition for TO PAR column (C) for under par (green)
under_par_rule = Rule(type="expression", formula=['LEFT(C2,1)="-"'], dxf=DifferentialStyle(fill=green_fill))
ws1.conditional_formatting.add('C2:C100', under_par_rule)

# Create condition for TO PAR column (C) for over par (red)
over_par_rule = Rule(type="expression", formula=['LEFT(C2,1)="+"'], dxf=DifferentialStyle(fill=red_fill))
ws1.conditional_formatting.add('C2:C100', over_par_rule)

# Apply same conditional formatting to TODAY column (E)
ws1.conditional_formatting.add('E2:E100', under_par_rule)
ws1.conditional_formatting.add('E2:E100', over_par_rule)

# Add a note about cut line in cell A1
ws1.cell(row=1, column=12).value = "Note: MC = Missed Cut"
ws1.cell(row=1, column=12).font = Font(italic=True)

# Format Tab 2: Participants Leaderboard
# Define columns for Tab 2
columns_tab2 = ["RANK", "PARTICIPANT", "TOTAL SCORE", "GROUPS SURVIVED", "STATUS", "PLAYER GROUP 1", 
                "PLAYER GROUP 2", "PLAYER GROUP 3", "PLAYER GROUP 4", "PLAYER GROUP 5", 
                "PLAYER GROUP 6", "PLAYER GROUP 7", "PLAYER GROUP 8", "PLAYER GROUP 9",
                "PLAYER GROUP 10", "PLAYER GROUP 11", "PLAYER GROUP 12"]

# Set column headers
for col_idx, col_name in enumerate(columns_tab2, 1):
    ws2.cell(row=1, column=col_idx).value = col_name
    ws2.cell(row=1, column=col_idx).font = Font(bold=True)
    ws2.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

# Adjust column widths for better visibility
ws2.column_dimensions['A'].width = 8   # RANK
ws2.column_dimensions['B'].width = 25  # PARTICIPANT
ws2.column_dimensions['C'].width = 15  # TOTAL SCORE
ws2.column_dimensions['D'].width = 18  # GROUPS SURVIVED
ws2.column_dimensions['E'].width = 15  # STATUS

# Set player group column widths
for i in range(6, 18):  # Columns F through R
    ws2.column_dimensions[get_column_letter(i)].width = 20

# Add sample participant data
sample_participants = [
    ["1", "John Smith", "-22", "9", "Active", "Scottie Scheffler", "Tommy Fleetwood", "Xander Schauffele", 
     "Justin Thomas", "Jon Rahm", "Hideki Matsuyama", "Max Homa", "Ludvig Åberg", "Patrick Cantlay", 
     "Collin Morikawa", "Jordan Spieth", "Rory McIlroy"],
    ["2", "Jane Doe", "-18", "8", "Active", "Collin Morikawa", "Max Homa", "Justin Thomas", 
     "Xander Schauffele", "Tommy Fleetwood", "Jon Rahm", "Ludvig Åberg", "Patrick Cantlay", 
     "Hideki Matsuyama", "Scottie Scheffler", "Jordan Spieth", "Rory McIlroy"],
    ["3", "Michael Johnson", "-15", "10", "Active", "Tommy Fleetwood", "Scottie Scheffler", "Jon Rahm", 
     "Hideki Matsuyama", "Max Homa", "Xander Schauffele", "Justin Thomas", "Ludvig Åberg", 
     "Patrick Cantlay", "Collin Morikawa", "Jordan Spieth", "Rory McIlroy"],
    ["4", "Sarah Williams", "-12", "7", "Active", "Scottie Scheffler", "Collin Morikawa", "Tommy Fleetwood", 
     "Max Homa", "Ludvig Åberg", "Jordan Spieth", "Rory McIlroy", "Xander Schauffele", 
     "Patrick Cantlay", "Hideki Matsuyama", "Justin Thomas", "Jon Rahm"],
    ["5", "Robert Brown", "-8", "6", "Eliminated", "Collin Morikawa", "Jordan Spieth", "Rory McIlroy", 
     "Tommy Fleetwood", "Scottie Scheffler", "Max Homa", "Ludvig Åberg", "Xander Schauffele", 
     "Patrick Cantlay", "Hideki Matsuyama", "Justin Thomas", "Jon Rahm"]
]

# Add the sample participant data to Tab 2
for row_idx, row_data in enumerate(sample_participants, 2):  # Start from row 2
    for col_idx, cell_value in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx).value = cell_value
        ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

# Apply conditional formatting to STATUS column
eliminated_rule = Rule(type="expression", formula=['E2="Eliminated"'], dxf=DifferentialStyle(fill=red_fill))
ws2.conditional_formatting.add('E2:E100', eliminated_rule)

active_rule = Rule(type="expression", formula=['E2="Active"'], dxf=DifferentialStyle(fill=green_fill))
ws2.conditional_formatting.add('E2:E100', active_rule)

# Add formulas for real implementation (commented examples)
# Add notes and formula explanation
ws2.cell(row=20, column=1).value = "Formula Implementation Instructions:"
ws2.cell(row=20, column=1).font = Font(bold=True)

ws2.cell(row=21, column=1).value = "1. GROUPS SURVIVED Formula (Column D):"
ws2.cell(row=21, column=1).font = Font(bold=True)
ws2.cell(row=22, column=1).value = "=COUNTIFS('CBS Leaderboard'!B:B,F2,'CBS Leaderboard'!K:K,\"<>CUT\") + COUNTIFS('CBS Leaderboard'!B:B,G2,'CBS Leaderboard'!K:K,\"<>CUT\") + ... (repeat for all player groups)"
ws2.cell(row=23, column=1).value = "This counts how many of the participant's selected players made the cut."

ws2.cell(row=25, column=1).value = "2. STATUS Formula (Column E):"
ws2.cell(row=25, column=1).font = Font(bold=True)
ws2.cell(row=26, column=1).value = "=IF(D2>=7,\"Active\",\"Eliminated\")"
ws2.cell(row=27, column=1).value = "This marks a participant as eliminated if fewer than 7 of their groups survived the cut."

ws2.cell(row=29, column=1).value = "3. TOTAL SCORE Formula (Column C):"
ws2.cell(row=29, column=1).font = Font(bold=True)
ws2.cell(row=30, column=1).value = "=IF(D2>=7,SUM(SMALL(IF(COUNTIF('CBS Leaderboard'!B:B,{F2,G2,H2,I2,J2,K2,L2,M2,N2,O2,P2,Q2})>0,VLOOKUP({F2,G2,H2,I2,J2,K2,L2,M2,N2,O2,P2,Q2},'CBS Leaderboard'!B:C,2,FALSE)),{1,2,3,4,5,6,7})),SUM(IF(COUNTIF('CBS Leaderboard'!B:B,{F2,G2,H2,I2,J2,K2,L2,M2,N2,O2,P2,Q2})>0,VLOOKUP({F2,G2,H2,I2,J2,K2,L2,M2,N2,O2,P2,Q2},'CBS Leaderboard'!B:C,2,FALSE))))"
ws2.cell(row=31, column=1).value = "This gets the TO PAR scores for each player, then either sums the best 7 scores (if 7+ groups survived) or sums all scores."

ws2.cell(row=33, column=1).value = "4. RANK Formula (Column A):"
ws2.cell(row=33, column=1).font = Font(bold=True)
ws2.cell(row=34, column=1).value = "=RANK.EQ(C2,$C$2:$C$100)"
ws2.cell(row=35, column=1).value = "This ranks participants by their total scores (lowest score is best)."

# Add data validation for CUT column in Tab 1
cut_validation = DataValidation(type="list", formula1='"CUT,"')
ws1.add_data_validation(cut_validation)
cut_validation.add('K2:K100')

# Save the workbook
wb.save(output_file)

# Output instructions for implementing and using the workbook
print(f"\nExcel workbook '{output_file}' created successfully with two tabs.")

print("\n==== IMPLEMENTATION INSTRUCTIONS ====")
print("\n1. TAB 1: CBS LEADERBOARD")
print("   - This tab mimics the CBS Sports Masters leaderboard format")
print("   - Update player data manually after each round or throughout the tournament")
print("   - To mark a player as having missed the cut, enter 'CUT' in the CUT column")
print("   - The TO PAR and TODAY columns will automatically color-code (red for over par, green for under par)")

print("\n2. TAB 2: PARTICIPANTS LEADERBOARD")
print("   - This tab tracks pool participants and their scores")
print("   - Enter participant names and their 12 selected player groups")
print("   - Formulas to implement (see notes at bottom of Tab 2):")
print("     a. Groups Survived: count players who made the cut")
print("     b. Status: 'Active' if 7+ groups survived, otherwise 'Eliminated'")
print("     c. Total Score: Sum of best 7 scores for active participants")
print("     d. Rank: Automatically ranks participants by total score")

print("\n3. DATA ENTRY & UPDATES")
print("   - Update Tab 1 after each round with latest scores")
print("   - Tab 2 will automatically calculate participant standings if proper formulas are implemented")
print("   - Add actual player names from the tournament to Tab 1")
print("   - Enter actual participant selections in Tab 2")

print("\n4. FORMULA IMPLEMENTATION")
print("   - The generated Excel file includes detailed formula instructions at the bottom of Tab 2")
print("   - Formulas need to be adapted based on final worksheet structure and cell references")
print("   - Array formulas may require using Ctrl+Shift+Enter in older Excel versions")

print("\n5. GITHUB UPLOAD")
print("   - Upload the final Excel file to https://github.com/rocketballer/SlatesGJG")
print("   - Include README.md with usage instructions if needed")

print("\nNOTE: This script created a template with sample data. You'll need to replace")
print("      with actual tournament data and participant selections.") 